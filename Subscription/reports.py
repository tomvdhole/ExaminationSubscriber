from django.db import models
from .models import Participant, Category
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font

class SubscriptionReport:
    def __init__(self):
        self.__column = 3
        self.__workbook = Workbook()

    def create_reports(self):
        self.__create_kids_report()
        self.__create_youth_report()
        self.__create_beginners_report()
        self.__create_advanced_report()
        self.__workbook.save(filename = 'Reports.xlsx')
        return self.__workbook

    #private methods
    def __create_kids_report(self):
        self.__create_kids_form_report()
        self.__create_kids_rhytm_report()
        self.__create_kids_official_report()

    def __create_youth_report(self):
        self.__create_youth_form_report()
        self.__create_youth_rhytm_report()
        self.__create_youth_official_report()

    def __create_beginners_report(self):
        self.__create_beginners_form_report()
        self.__create_beginners_rhytm_report()
        self.__create_beginners_official_report()

    def __create_advanced_report(self):
        self.__create_advanced_form_report()
        self.__create_advanced_rhytm_report()
        self.__create_advanced_official_report()

    def __create_kids_form_report(self):
        self.__ws = self.__workbook.active
        self.__ws.title = 'Kids Form'
        self.__make_header('Kids Form')
        kids = Participant.objects.filter(category__type_of_participants__exact= 'Kids', category__examination_type__exact = 'Form')
        self.__fill_report(kids)

    def __create_youth_form_report(self):
        self.__ws = self.__workbook.create_sheet(title="Youth Form")
        self.__make_header('Youth Form')
        youth = Participant.objects.filter(category__type_of_participants__exact= 'Youth', category__examination_type__exact = 'Form')
        self.__fill_report(youth)

    def __create_beginners_form_report(self):
        self.__ws = self.__workbook.create_sheet(title="Beginners Form")
        self.__make_header('Beginners Form')
        beginners = Participant.objects.filter(category__type_of_participants__exact= 'Beginners', category__examination_type__exact = 'Form')
        self.__fill_report(beginners)

    def __create_advanced_form_report(self):
        self.__ws = self.__workbook.create_sheet(title="Advanced Form")
        self.__make_header('Advanced Form')
        advanced = Participant.objects.filter(category__type_of_participants__exact= 'Advanced', category__examination_type__exact = 'Form')
        self.__fill_report(advanced)

    def __create_kids_rhytm_report(self):
        self.__ws = self.__workbook.create_sheet(title="Kids Rhytm")
        self.__make_header('Kids Rhytm')
        kids = Participant.objects.filter(category__type_of_participants__exact= 'Kids', category__examination_type__exact = 'Rhytm')
        self.__fill_report(kids)

    def __create_youth_rhytm_report(self):
        self.__ws = self.__workbook.create_sheet(title="Youth Rhytm")
        self.__make_header('Youth Rhytm')
        youth = Participant.objects.filter(category__type_of_participants__exact= 'Youth', category__examination_type__exact = 'Rhytm')
        self.__fill_report(youth)

    def __create_beginners_rhytm_report(self):
        self.__ws = self.__workbook.create_sheet(title="Beginners Rhytm")
        self.__make_header('Beginners Rhytm')
        beginners = Participant.objects.filter(category__type_of_participants__exact= 'Beginners', category__examination_type__exact = 'Rhytm')
        self.__fill_report(beginners)

    def __create_advanced_rhytm_report(self):
        self.__ws = self.__workbook.create_sheet(title="Advanced Rhytm")
        self.__make_header('Advanced Rhytm')
        advanced = Participant.objects.filter(category__type_of_participants__exact= 'Advanced', category__examination_type__exact = 'Rhytm')
        self.__fill_report(advanced)

    def __create_kids_official_report(self):
        self.__ws = self.__workbook.create_sheet(title="Kids Official")
        self.__make_header('Kids Official')
        kids = Participant.objects.filter(category__type_of_participants__exact= 'Kids', category__examination_type__exact = 'Official')
        self.__fill_report(kids)

    def __create_youth_official_report(self):
        self.__ws = self.__workbook.create_sheet(title="Youth Official")
        self.__make_header('Youth Official')
        youth = Participant.objects.filter(category__type_of_participants__exact= 'Youth', category__examination_type__exact = 'Official')
        self.__fill_report(youth)

    def __create_beginners_official_report(self):
        self.__ws = self.__workbook.create_sheet(title="Beginners Official")
        self.__make_header('Beginners Official')
        beginners = Participant.objects.filter(category__type_of_participants__exact= 'Beginners', category__examination_type__exact = 'Official')
        self.__fill_report(beginners)

    def __create_advanced_official_report(self):
        self.__ws = self.__workbook.create_sheet(title="Advanced Official")
        self.__make_header('Advanced Official')
        advanced = Participant.objects.filter(category__type_of_participants__exact= 'Advanced', category__examination_type__exact = 'Official')
        self.__fill_report(advanced)

    def __make_header(self, title):
        self.__ws.add_image(Image('Subscription/static/Subscription/images/Taisho.gif'), 'A1')
        self.__ws['C1'] = title
        self.__ws['C1'].font = Font(bold = True)
        self.__ws.cell(column = 1, row = 10, value = 'Naam').font = Font(bold = True)
        self.__ws.cell(column = 2, row = 10, value = 'Voornaam').font = Font(bold = True)
        self.__ws.cell(column = 3, row = 10, value = 'Leeftijd').font = Font(bold = True)
        self.__ws.cell(column = 4, row = 10, value = 'Graad').font = Font(bold = True)
        self.__ws.cell(column = 5, row = 10, value = 'Aantal rode streepjes').font = Font(bold = True)
        self.__ws.cell(column = 6, row = 10, value = 'Licentie nummer').font = Font(bold = True)
        self.__ws.cell(column = 7, row = 10, value = 'Licentie geldig tot ').font = Font(bold = True)
        self.__ws.cell(column = 8, row = 10, value = 'Gsm').font = Font(bold = True)
        self.__ws.cell(column = 9, row = 10, value = 'Email').font = Font(bold = True)

    def __fill_report(self, participants):
        self.__column = 1
        self.__row = 11
        for participant in participants:
            self.__ws.cell(column = self.__column, row = self.__row, value = participant.last_name)
            self.__column += 1
            self.__ws.cell(column = self.__column, row = self.__row, value = participant.first_name)
            self.__column += 1
            self.__ws.cell(column = self.__column, row = self.__row, value = participant.age)
            self.__column += 1
            self.__ws.cell(column = self.__column, row = self.__row, value = participant.grade)
            self.__column += 1
            self.__ws.cell(column = self.__column, row = self.__row, value = participant.number_of_red_ribbons)
            self.__column += 1
            self.__ws.cell(column = self.__column, row = self.__row, value = participant.license_number)
            self.__column += 1
            self.__ws.cell(column = self.__column, row = self.__row, value = participant.date_till_license_is_valid)
            self.__column += 1
            self.__ws.cell(column = self.__column, row = self.__row, value = participant.gsm)
            self.__column += 1
            self.__ws.cell(column = self.__column, row = self.__row, value = participant.email)
            self.__row += 1
            self.__column = 1
